
import os
from math import isnan
import datetime as dt
import shutil
import dateutil.parser as dateparser
from pathlib import Path
from subprocess import Popen, PIPE, run

import openpyxl as pyxl
import pandas as pd
from pandas.core.frame import DataFrame

import config
from crypt_utilities.hashes import derive
from crypt_utilities.asymmetric import rsa_encrypt, RSAPublicKey, RSAPrivateKey, rsa_decrypt
from editor.protected_data import (
    hash_and_save_encrypted, get_encrypted, hash_sheet_ids_and_save, get_sheet_hashed_ids,
    num_iters_hashes
)
from editor.protected_data import pd_file_path

data_dir_name = './.data'
backup_dir = './.backup'
main_execution_path = Path(__file__).resolve().parent
data_dir_path = main_execution_path/data_dir_name

date_format = config.get('date_format', else_return='%d/%m/%Y')

def get_excel_path() -> Path:
    excel_path_str = config.get('excel_path')
    if excel_path_str != None: 
        excel_path = Path(excel_path_str)
        if not excel_path.is_absolute():
            excel_path = main_execution_path/excel_path_str
        
    return excel_path

def launch() -> bool:
    process = run(f'start excel "{get_excel_path()}"', shell=True, cwd=os.getcwd(), stdout=PIPE, stderr=PIPE)
    if process.returncode == 0:
        return True
    return False
    
def refresh_csv_sheets() -> dict:
    print("[%] Refreshing csv sheets...")
    if empty(): return
    excel_path = get_excel_path()
    excel = pd.ExcelFile(excel_path)
    sheet_names = excel.sheet_names; sheets_info = {}
    # Eliminamos las que ya no exitan en el excel (cambio de nombre o eliminadas)
    existing_sh_names = list_csv_sheets()
    for ex_sh in existing_sh_names:
        if ex_sh not in sheet_names:
            os.remove(data_dir_path/(ex_sh+".csv"))
    
    for sheet_name in sheet_names:
        sheet:DataFrame = excel.parse(sheet_name)
        sheet_dest_path = data_dir_path/(sheet_name+".csv")
        sheets_info[sheet_name] = sheet_dest_path
        sheet.to_csv(sheet_dest_path, index=False, date_format=date_format)
        sheet.columns
        
    return sheets_info

def list_csv_sheets() -> list:
    filtered = filter(lambda f: f.endswith('.csv'), os.listdir(data_dir_path))
    return list(map(lambda f: f.removesuffix('.csv'), filtered))

def check_id_value(id_value:str, sheet_n:str=None) -> dict:
    id_field = config.get('id_field')
    sfields:list = config.get('sensitive_fields')
    if id_field is None:
        raise Exception("El campo 'id_field' en .config.json esta vacio, no se ha especificado un campo identificador")
    if empty(): return {}
    data_types:list = config.get('data_types')
    excel_path = get_excel_path()
    excel = pd.ExcelFile(excel_path)
    sheet_names = excel.sheet_names; found = {}
    dtype = None
    if id_field in data_types:
        dtype = data_types.get(id_field, None)
    if dtype is not None:
        id_value = _parse_elem(id_value, dtype)
    for sheet_name in sheet_names:
        if sheet_n is not None and sheet_n != sheet_name: continue
        if id_field in sfields and os.path.exists(pd_file_path):
            array:list = get_sheet_hashed_ids(sheet_name=sheet_name, decrypt_hashes=True)
            if array is not None:
                for index, hashed_id in enumerate(array):
                    for hash_id_val, salt in hashed_id.items():
                        hashed_val = derive(str(id_value).encode(), salt, iterations=num_iters_hashes)
                        if hashed_val == hash_id_val:
                            found[sheet_name] = index + 1
                            break
                    else:
                        continue
                    break
        else:
            df = excel.parse(sheet_name)
            csv_dict = df.to_dict("list")
            if id_field in csv_dict:
                array = csv_dict[id_field]
                if dtype is not None:
                    for i, a in enumerate(array):
                        array[i] = _parse_elem(a, dtype)
                if id_value in array:
                    found[sheet_name] = array.index(id_value) + 1
                    break
    return found
        
def update_excel():
    print("[%] Updating excel from csv sheets...")
    if empty(): return
    csv_sheet_names = list_csv_sheets()
    sheets_dict = {}
    for sheet_name in csv_sheet_names:
        df = pd.read_csv(data_dir_path/(sheet_name+".csv"))
        csv_dict = df.to_dict("list")
        sheets_dict[sheet_name] = csv_dict
    wb = pyxl.load_workbook(get_excel_path())
    _write_dict_into_wb(wb, sheets_dict, encrypted=True)
            
def _parse_elem(elem:str, dtype:str) -> any:
    try:
        if dtype == 'int':
            elem:float = float(elem)
            if elem.is_integer():
                return int(elem)
        elif dtype == 'float':
            return float(elem)
        elif dtype == 'str-q':
            if not str(elem).startswith("'"):
                return "'"+str(elem)
        elif dtype == 'date':
            raw_date_str = elem.split(" ")[0]
            if "/" in raw_date_str: sep = "/"
            elif "-" in raw_date_str: sep = "-"
            first_pos = raw_date_str.split(sep)[0]
            if len(first_pos) == 4:
                date = dt.datetime.strptime(raw_date_str, f"%Y{sep}%m{sep}%d")
            else:
                date = dt.datetime.strptime(raw_date_str, f"%d{sep}%m{sep}%Y")
            date_str = str(date.strftime(date_format))
            return str(date_str)
    except Exception as err:
        pass
    return str(elem) 

def isempty(elem:any) -> bool:
    is_nan = False
    try: is_nan = isnan(elem)
    except: pass
    return elem == "" or elem is None or is_nan

def _write_dict_into_wb(wb:pyxl.Workbook, sheets_dict:dict, encrypted=False):
    print("[%] Writing into excel sheets...")
    data_types:list = config.get('data_types')
    sfields:list = config.get('sensitive_fields')
    if sfields is None: sfields = []
    # Guardar por celdas en cada ws  
    excel_path = get_excel_path()
    wb = pyxl.load_workbook(excel_path)
    for sheet_name, sheet_info in sheets_dict.items():
        ws = wb[sheet_name]
        for colum, header in enumerate(sheet_info.keys(),1):
            dtype = None
            if data_types is not None and header in data_types:
                dtype = data_types[header]
            colum_elements = sheet_info[header]
            ws.cell(1, colum).value = header
            for row, elem  in enumerate(colum_elements, 2):
                if not encrypted or not header in sfields:
                    if isempty(elem): 
                        elem = ""
                    elif dtype is not None:
                        elem = _parse_elem(elem, dtype)
                ws.cell(row, colum).value = elem
    wb.save(excel_path)
             
def backup():
    print("[%] Executing excel backup...")
    not_protected = False
    if not os.path.exists(pd_file_path):
        not_protected = True
        bkup = config.get('backup-not-encrypted')
        if not bkup:
            print("[!] Excel not encrypted, aborting backup...")
            return
    try:
        backup_path = data_dir_path/backup_dir
        if not os.path.exists(backup_path):
            os.mkdir(backup_path)
        with open(get_excel_path(), 'rb') as file:
            content = file.read()
        date_str = _get_date(path_friendly=True)
        if not_protected: date_str = "NP-"+date_str
        dir_path = backup_path/date_str
        os.mkdir(dir_path)
        bname = date_str+".xlsx"
        with open(dir_path/bname, 'wb') as file:
            file.write(content)
        shutil.copy(pd_file_path, dir_path/pd_file_path.name)
    except FileNotFoundError: pass

def decrypt_excel(private_key:RSAPrivateKey):
    print("[%] Decrypting excel fields...")
    if empty(): return
    sfields = config.get('sensitive_fields')
    excel_path = get_excel_path()
    excel = pd.ExcelFile(excel_path)
    sheet_names = excel.sheet_names
    decrypted_sheets = {}
    for sheet_name in sheet_names:
        sheet:DataFrame = excel.parse(sheet_name)
        csv_dict = sheet.to_dict("list")
        decrypted_sheet = {}
        for header in csv_dict:
            colum = csv_dict[header]
            decrypted_colum = []
            for elem in colum:
                if isempty(elem): elem = ""
                elem = str(elem)
                ciphertext = get_encrypted(elem.encode()); 
                if ciphertext is not None:
                    elem = rsa_decrypt(ciphertext, private_key).decode()
                decrypted_colum.append(elem)
            colum = decrypted_colum
            decrypted_sheet[header] = colum
        decrypted_sheets[sheet_name] = decrypted_sheet
    
    if os.path.exists(pd_file_path):
        os.remove(pd_file_path)
    wb = pyxl.load_workbook(excel_path)
    _write_dict_into_wb(wb, decrypted_sheets)
       

def protect_sensitive_data(public_key:RSAPublicKey):
    print("[%] Encrypting excel data...")
    if empty(): return
    sfields:list = config.get('sensitive_fields')
    if sfields is None: sfields = []
    excel = pd.ExcelFile(get_excel_path())
    sheet_names = excel.sheet_names
    protected_sheets = {}
    for sheet_name in sheet_names:
        sheet:DataFrame = excel.parse(sheet_name)
        csv_dict = sheet.to_dict("list")
        protected_sheet = {}
        for header in csv_dict:
            colum = csv_dict[header]
            # Encriptamos datos sensibles
            if header in sfields:
                if config.is_id_field(header):
                    hash_sheet_ids_and_save(sheet_name, colum)
                protected_colum = []
                for elem in colum:
                    if isempty(elem): 
                        elem = ""
                    else:
                        elem = str(elem)
                        encrypted = get_encrypted(elem.encode())
                        if encrypted is None:
                            elem = str(elem)
                            ciphertext = rsa_encrypt(elem.encode(), public_key)
                            salted_hash = hash_and_save_encrypted(ciphertext)
                            elem = salted_hash.decode()
                    protected_colum.append(elem)
                colum = protected_colum
            protected_sheet[header] = colum
        protected_sheets[sheet_name] = protected_sheet
    # Guardar por celdas en cada ws  
    wb = pyxl.load_workbook(get_excel_path())
    _write_dict_into_wb(wb, protected_sheets, encrypted=True)
        
def _get_date(path_friendly:bool=False) -> str:
    datetime = dt.datetime.now()
    if path_friendly:
        date = str(datetime.date())
        time = str(datetime.time()).replace(':', "-").replace('.', '_')
        return date+"_"+time
    else:
        return str(datetime)
        
def check():
    try:
        open(get_excel_path(), 'r+')
    except IOError:
        # El excel sigue abierto
        return False
    else:
        return True
    
def empty():
    df = pd.read_excel(get_excel_path())
    return df.empty
    
if __name__ == "__main__":
    refresh_csv_sheets()